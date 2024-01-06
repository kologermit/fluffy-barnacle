from openpyxl import load_workbook
import logging

from aiogram import types
from aiogram.dispatcher import FSMContext

from loader import dp
from keyboards import LoadSheets
from db import *
from state import *


@dp.callback_query_handler(text='load_sheets')
async def load_sheets(c: types.CallbackQuery):
    await c.message.answer('Хорошо, выберите таблицу:', reply_markup=LoadSheets.ikb)
    await c.answer()


@dp.message_handler(state=TypePersonal_Money_State.stating, content_types=types.ContentTypes.DOCUMENT)
async def desc_of_type_persons_load(m: types.Message, state: FSMContext):
    load_message = await m.answer("Началась загрузка базы сообщений...")
    try:
        await TypePersonal_Money.all().delete()
        file = await dp.bot.get_file(m.document.file_id)
        await dp.bot.download_file(file.file_path, "menu.xlsx")
        wb = load_workbook("menu.xlsx")
        sheet = wb.active
        for cells in sheet.iter_rows():
            item = [cell for cell in cells]
            key = item[0].value
            text = item[1].value
            home_work = item[2].value
            congratulation = item[3].value
            await TypePersonal_Money.create(key=key, description=text, home_work=home_work, congratulation=congratulation)
        await dp.bot.edit_message_text(
            "Отлично база сообщений загружена!",
            chat_id=m.chat.id,
            message_id=load_message.message_id
        )
    except Exception as e:
        await m.answer(e)
        await dp.bot.edit_message_text(
            "При загрузке базы сообщений произошла ошибка. Проверьте структуру файла и повторите попытку.",
            chat_id=m.chat.id,
            message_id=load_message.message_id
        )
    await state.finish()


@dp.callback_query_handler(text='desc_of_authory_in_business')
async def desc_of_authory_in_business(c: types.CallbackQuery, state: FSMContext):
    await c.message.answer('Отлично, отправьте нужную таблицу:')
    await state.set_state(AuthorityInBusiness_Money_State.stating.state)
    await c.answer()


@dp.message_handler(state=AuthorityInBusiness_Money_State.stating, content_types=types.ContentTypes.DOCUMENT)
async def desc_of_authory_in_business_load(m: types.Message, state: FSMContext):
    load_message = await m.answer("Началась загрузка базы сообщений...")
    try:
        await AuthorityInBusiness_Money.all().delete()
        file = await dp.bot.get_file(m.document.file_id)
        await dp.bot.download_file(file.file_path, "menu.xlsx")
        wb = load_workbook("menu.xlsx")
        sheet = wb.active
        for cells in sheet.iter_rows():
            item = [cell for cell in cells]
            key = item[0].value
            text = item[1].value
            home_work = item[2].value
            congratulation = item[3].value
            await AuthorityInBusiness_Money.create(key=key, description=text, home_work=home_work, congratulation=congratulation)
        await dp.bot.edit_message_text(
            "Отлично база сообщений загружена!",
            chat_id=m.chat.id,
            message_id=load_message.message_id
        )
    except Exception as e:
        await m.answer(e)
        await dp.bot.edit_message_text(
            "При загрузке базы сообщений произошла ошибка. Проверьте структуру файла и повторите попытку.",
            chat_id=m.chat.id,
            message_id=load_message.message_id
        )
    await state.finish()


@dp.callback_query_handler(text='desc_of_strategy_profiles')
async def desc_of_strategy_profiles(c: types.CallbackQuery, state: FSMContext):
    await c.message.answer('Отлично, отправьте нужную таблицу:')
    await state.set_state(StrategyProfiles_Money_State.stating.state)
    await c.answer()

@dp.callback_query_handler(text='desc_of_products')
async def desc_of_products(c: types.CallbackQuery, state: FSMContext):
    await c.message.answer('Отлично, отправьте нужную таблицу:')
    await state.set_state(Products_State.stating.state)
    await c.answer()

@dp.message_handler(state=StrategyProfiles_Money_State.stating, content_types=types.ContentTypes.DOCUMENT)
async def strategy_profile(m: types.Message, state: FSMContext):
    load_message = await m.answer("Началась загрузка базы сообщений...")
    try:
        await StrategyProfiles_Money.all().delete()
        file = await dp.bot.get_file(m.document.file_id)
        await dp.bot.download_file(file.file_path, "menu.xlsx")
        wb = load_workbook("menu.xlsx")
        sheet = wb.active
        for cells in sheet.iter_rows():
            items = [("-" if cell.value is None else cell.value) for cell in cells]
            logging.info(items)
            if len(items) < 3:
                items.append("-")
            key, name, description, home_work, congratulation = items
            await StrategyProfiles_Money.create(key=key, description=description, home_work=home_work, congratulation=congratulation, name=name)
            print(key)
            print(name)
            print(description)
            print(home_work)
            print(congratulation)
        await dp.bot.edit_message_text(
            "Отлично база сообщений загружена!",
            chat_id=m.chat.id,
            message_id=load_message.message_id
        )
    except Exception as e:
        await m.answer(e)
        await dp.bot.edit_message_text(
            "При загрузке базы сообщений произошла ошибка. Проверьте структуру файла и повторите попытку.",
            chat_id=m.chat.id,
            message_id=load_message.message_id
        )
    await state.finish()


@dp.callback_query_handler(text='desc_of_type_persons')
async def desc_of_type_persons(c: types.CallbackQuery, state: FSMContext):
    await c.message.answer('Отлично, отправьте нужную таблицу:')
    await state.set_state(TypePersonal_Money_State.stating.state)
    await c.answer()

@dp.message_handler(state=Products_State.stating, content_types=types.ContentTypes.DOCUMENT)
async def strategy_profile(m: types.Message, state: FSMContext):
    load_message = await m.answer("Началась загрузка базы сообщений...")
    try:
        await Products.all().delete()
        file = await dp.bot.get_file(m.document.file_id)
        await dp.bot.download_file(file.file_path, "menu.xlsx")
        wb = load_workbook("menu.xlsx")
        sheet = wb.active
        for cells in sheet.iter_rows():
            items = [("-" if cell.value is None else cell.value) for cell in cells]
            logging.info(items)
            if len(items) < 3:
                items.append("-")
            name, price, description = items
            await Products.create(name=name, price=price, description=description)
            print(name)
            print(price)
            print(description)
        await dp.bot.edit_message_text(
            "Отлично база сообщений загружена!",
            chat_id=m.chat.id,
            message_id=load_message.message_id
        )
    except Exception as e:
        logging.exception(e)
        await m.answer(e)
        await dp.bot.edit_message_text(
            "При загрузке базы сообщений произошла ошибка. Проверьте структуру файла и повторите попытку.",
            chat_id=m.chat.id,
            message_id=load_message.message_id
        )
    await state.finish()