from aiogram import types
from openpyxl import * 
from openpyxl.styles import *
import logging
from db import *
from data.config import admin_id
from loader import dp
from keyboards import AdminMenu, TypesBaseRegistration
from prodamus.prodamus import *

@dp.message_handler(text='Показать список клиентов')
async def show_list_clients(m: types.Message):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Клиенты'
    r_s = await AuthorityInBusiness_Money.filter().all()
    ws.append(["1", "2", "3", "4", "5"])
    ft = Font(
        name='Calibri',
        size=11,
        bold=True,
    )
    ws["A1"].font = ft
    ws["B1"].font = ft
    ws["C1"].font = ft
    ws["D1"].font = ft
    ws["E1"].font = ft
    for item in r_s:
        id = item.id
        key = item.key
        description = item.description
        home_work = item.home_work
        congratulation = item.congratulation
        print(id)
        print(key)
        print(description)
        print(home_work)
        print(congratulation)

@dp.message_handler(commands=['admin'])
async def admin_start(m: types.Message):
    if m.from_user.id in admin_id:
        await m.answer('Добрый день!\n'
                       'Выберите действия:', reply_markup=AdminMenu.ikb)

@dp.message_handler(commands=['test'])
async def test(m: types.Message):
    logging.info(prodamus_create_url({
        "name": "Имя товара",
        "price": 100,
        "quantity": 1,
        "sku": m.from_user.id
    }, m.from_user.id))

@dp.callback_query_handler(text='check_recs_regs')
async def check_recs_regs(c: types.CallbackQuery):
    await c.message.answer('Хорошо, выберите категорию', reply_markup=TypesBaseRegistration.ikb)

@dp.callback_query_handler(text_startswith='type_reg:')
async def choice_sphere(c: types.CallbackQuery):
    sphere_reg = c.data[9:]
    r_s = await BaseRegistration.filter(sphere=sphere_reg).all()
    for item in r_s:
        await c.message.answer(f'<b>Короткая ссылка:</b> @{item.tg_un_user}\n'
                               f'<b>Сфера:</b> {item.sphere}\n'
                               f'<b>Дата рождения:</b> {item.born_date}\n'
                               f'<b>Время рождения:</b> {item.born_time}\n'
                               f'<b>Город рождения:</b> {item.born_city}')

